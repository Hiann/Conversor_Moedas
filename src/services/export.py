"""Serviço de exportação de dados."""

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from src.core.models import Conversao, ExportacaoConfig


class ExportService:
    """Serviço para exportar dados em vários formatos."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    def exportar(
        self, 
        conversoes: List[Conversao], 
        config: ExportacaoConfig
    ) -> Path:
        """
        Exporta conversões no formato especificado.
        
        Args:
            conversoes: Lista de conversões
            config: Configuração de exportação
            
        Returns:
            Caminho do arquivo gerado
        """
        arquivo = Path(config.arquivo)
        arquivo.parent.mkdir(parents=True, exist_ok=True)
        
        if config.formato == "excel":
            return self._exportar_excel(conversoes, arquivo)
        elif config.formato == "csv":
            return self._exportar_csv(conversoes, arquivo)
        elif config.formato == "json":
            return self._exportar_json(conversoes, arquivo)
        elif config.formato == "pdf":
            return self._exportar_pdf(conversoes, arquivo)
        else:
            raise ValueError(f"Formato não suportado: {config.formato}")
    
    def _exportar_excel(self, conversoes: List[Conversao], arquivo: Path) -> Path:
        """Exporta para Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversões"
        
        # Cabeçalhos
        headers = ["ID", "Data", "Origem", "Valor Origem", "Destino", "Valor Destino", "Taxa", "Notas"]
        ws.append(headers)
        
        # Estilo do cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for conv in conversoes:
            ws.append([
                conv.id,
                conv.timestamp.strftime("%d/%m/%Y %H:%M") if conv.timestamp else "",
                conv.moeda_origem,
                float(conv.valor_original),
                conv.moeda_destino,
                float(conv.valor_convertido),
                float(conv.taxa),
                conv.notas or ""
            ])
        
        # Ajusta larguras
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30
        
        wb.save(arquivo)
        return arquivo
    
    def _exportar_csv(self, conversoes: List[Conversao], arquivo: Path) -> Path:
        """Exporta para CSV."""
        with open(arquivo, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Cabeçalhos
            writer.writerow(["ID", "Data", "Origem", "Valor Origem", "Destino", "Valor Destino", "Taxa", "Notas"])
            
            # Dados
            for conv in conversoes:
                writer.writerow([
                    conv.id,
                    conv.timestamp.isoformat() if conv.timestamp else "",
                    conv.moeda_origem,
                    float(conv.valor_original),
                    conv.moeda_destino,
                    float(conv.valor_convertido),
                    float(conv.taxa),
                    conv.notas or ""
                ])
        
        return arquivo
    
    def _exportar_json(self, conversoes: List[Conversao], arquivo: Path) -> Path:
        """Exporta para JSON."""
        data = {
            "exportado_em": datetime.now().isoformat(),
            "total": len(conversoes),
            "conversoes": [
                {
                    "id": c.id,
                    "valor_original": float(c.valor_original),
                    "valor_convertido": float(c.valor_convertido),
                    "moeda_origem": c.moeda_origem,
                    "moeda_destino": c.moeda_destino,
                    "taxa": float(c.taxa),
                    "taxa_inversa": float(c.taxa_inversa),
                    "timestamp": c.timestamp.isoformat() if c.timestamp else None,
                    "notas": c.notas
                }
                for c in conversoes
            ]
        }
        
        with open(arquivo, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        return arquivo
    
    def _exportar_pdf(self, conversoes: List[Conversao], arquivo: Path) -> Path:
        """Exporta para PDF."""
        doc = SimpleDocTemplate(
            str(arquivo),
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=20
        )
        elements.append(Paragraph("Histórico de Conversões", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Data
        elements.append(Paragraph(
            f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            self.styles['Normal']
        ))
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabela
        data = [["Data", "Origem", "Valor", "Destino", "Valor", "Taxa"]]
        
        for conv in conversoes:
            data.append([
                conv.timestamp.strftime("%d/%m/%Y") if conv.timestamp else "",
                conv.moeda_origem,
                f"{float(conv.valor_original):,.2f}",
                conv.moeda_destino,
                f"{float(conv.valor_convertido):,.2f}",
                f"{float(conv.taxa):,.4f}"
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return arquivo
